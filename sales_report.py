import os
import re
import tempfile
import smtplib
from datetime import datetime, timezone, timedelta, date
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import openpyxl
import anthropic
from playwright.sync_api import sync_playwright


KST = timezone(timedelta(hours=9))
LOGIN_URL = "https://www.matetech.co.kr/login"
CHANNEL_URL = "https://www.matetech.co.kr/service/saa0060/sales/analysis/order-path"
STORE_URL = "https://www.matetech.co.kr/service/saa0010/sales/analysis/term"


def load_accounts():
    """환경변수에서 계정 목록 로드.

    MATETECH_USERNAME / MATETECH_PASSWORD / MATETECH_BRAND (1번 계정)
    MATETECH_USERNAME_2 / MATETECH_PASSWORD_2 / MATETECH_BRAND_2 (2번 계정)
    ... 계정 개수만큼 N을 늘려서 추가
    """
    accounts = []
    u = os.environ.get("MATETECH_USERNAME", "").strip()
    p = os.environ.get("MATETECH_PASSWORD", "").strip()
    if u and p:
        accounts.append({
            "username": u,
            "password": p,
            "brand": os.environ.get("MATETECH_BRAND", "").strip() or "브랜드 1",
        })
    i = 2
    while True:
        u = os.environ.get(f"MATETECH_USERNAME_{i}", "").strip()
        p = os.environ.get(f"MATETECH_PASSWORD_{i}", "").strip()
        if not (u and p):
            break
        accounts.append({
            "username": u,
            "password": p,
            "brand": os.environ.get(f"MATETECH_BRAND_{i}", "").strip() or f"브랜드 {i}",
        })
        i += 1
    return accounts


def get_date_ranges(report_type: str):
    """현재/비교 기간의 (start, end) 튜플 두 개 반환.

    daily: (어제, 어제) vs (그제, 그제)
    weekly: (어제-6, 어제) vs (어제-13, 어제-7)
    """
    today = datetime.now(KST).date()
    yesterday = today - timedelta(days=1)

    if report_type == "daily":
        cur = (yesterday, yesterday)
        prev = (yesterday - timedelta(days=1), yesterday - timedelta(days=1))
    else:
        cur_end = yesterday
        cur_start = yesterday - timedelta(days=6)
        prev_end = cur_start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=6)
        cur = (cur_start, cur_end)
        prev = (prev_start, prev_end)
    return cur, prev


def _query_and_download(page, start: date, end: date) -> str:
    """현재 페이지에서 지정 기간 입력 → 검색 → 엑셀 다운로드."""
    start_str = start.strftime("%Y%m%d")
    end_str = end.strftime("%Y%m%d")

    page.evaluate(
        """
        ({start, end}) => {
            const inputs = document.querySelectorAll('input.hasDatepicker');
            if (inputs.length < 2) {
                throw new Error('datepicker input not found: ' + inputs.length);
            }
            inputs.forEach(i => i.removeAttribute('readonly'));
            inputs[0].value = start;
            inputs[1].value = end;
            ['input', 'change', 'blur'].forEach(ev => {
                inputs.forEach(i => i.dispatchEvent(new Event(ev, {bubbles: true})));
            });
            if (window.jQuery) {
                window.jQuery(inputs).trigger('change').trigger('blur');
            }
        }
        """,
        {"start": start_str, "end": end_str},
    )
    page.wait_for_timeout(500)

    page.get_by_role("button", name="검색", exact=True).click()
    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(2000)

    with page.expect_download() as dl_info:
        page.get_by_role("button", name="엑셀 다운로드", exact=True).click()
    download = dl_info.value

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    download.save_as(tmp.name)
    return tmp.name


def _check_store_split(page):
    """기간별 매출분석 페이지의 '매장별 분리 (일자기준)' 체크박스 활성화."""
    try:
        page.evaluate(
            """
            () => {
                const labels = document.querySelectorAll('label, span, div');
                for (const el of labels) {
                    const txt = (el.textContent || '').trim();
                    if (txt === '매장별 분리 (일자기준)' || txt.startsWith('매장별 분리 (일자기준)')) {
                        const cb = el.querySelector('input[type=checkbox]')
                                || (el.previousElementSibling && el.previousElementSibling.querySelector && el.previousElementSibling.querySelector('input[type=checkbox]'))
                                || (el.parentElement && el.parentElement.querySelector('input[type=checkbox]'));
                        if (cb && !cb.checked) {
                            cb.click();
                            cb.dispatchEvent(new Event('change', {bubbles: true}));
                        }
                        return 'checked';
                    }
                }
                // fallback: try clicking first checkbox in the period section
                return 'not_found';
            }
            """
        )
    except Exception as e:
        print(f"  매장별 분리 체크 실패 (계속 진행): {e}")


def fetch_sales_data(username: str, password: str, report_type: str):
    """한 번 로그인으로 채널별/매장별 × 현재/비교 = 엑셀 4개 다운로드.

    Returns: (cur_channel, prev_channel, cur_store, prev_store, cur_range, prev_range)
    """
    cur_range, prev_range = get_date_ranges(report_type)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()

        page.goto(LOGIN_URL)
        page.fill("#member-id", username)
        page.fill("#member-pw", password)
        page.click("#member-btn")
        page.wait_for_load_state("networkidle")

        # 1) 채널별 매출 (주문경로별)
        page.goto(CHANNEL_URL)
        page.wait_for_load_state("networkidle")

        print(f"  [채널] 현재 기간 조회: {cur_range[0]} ~ {cur_range[1]}")
        cur_channel = _query_and_download(page, cur_range[0], cur_range[1])

        print(f"  [채널] 비교 기간 조회: {prev_range[0]} ~ {prev_range[1]}")
        prev_channel = _query_and_download(page, prev_range[0], prev_range[1])

        # 2) 매장별 매출 (기간별 매출분석)
        page.goto(STORE_URL)
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(1000)
        _check_store_split(page)
        page.wait_for_timeout(300)

        print(f"  [매장] 현재 기간 조회: {cur_range[0]} ~ {cur_range[1]}")
        cur_store = _query_and_download(page, cur_range[0], cur_range[1])

        print(f"  [매장] 비교 기간 조회: {prev_range[0]} ~ {prev_range[1]}")
        prev_store = _query_and_download(page, prev_range[0], prev_range[1])

        browser.close()

    return cur_channel, prev_channel, cur_store, prev_store, cur_range, prev_range


def parse_excel(filepath: str):
    """주문경로별 매출분석 엑셀. 4행 이후가 데이터 (1~2행 헤더, 3행 합계 라벨)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        rows.append({
            "order_type": row[0],
            "channel": row[1],
            "channel_detail": row[2],
            "count": int(row[3] or 0),
            "amount": int(row[5] or 0),
        })
    return rows


def parse_store_excel(filepath: str):
    """기간별 매장별 분리 엑셀. 4행 이후가 데이터.
    컬럼: 분기/월/일자, 일자, 매장코드, 매장명, 건수, 실매출액, ...
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 6:
            continue
        store_code = row[2]
        store_name = row[3]
        if not store_code or not store_name:
            continue
        rows.append({
            "store_code": str(store_code).strip(),
            "store_name": str(store_name).strip(),
            "count": int(row[4] or 0),
            "amount": int(row[5] or 0),
        })
    return rows


def aggregate_stores(rows):
    """매장코드 기준으로 합산 (여러 일자 누적). Returns: {code: {name, count, amount}}"""
    by_store = {}
    for r in rows:
        code = r["store_code"]
        if code not in by_store:
            by_store[code] = {
                "store_name": r["store_name"],
                "count": 0,
                "amount": 0,
            }
        by_store[code]["count"] += r["count"]
        by_store[code]["amount"] += r["amount"]
    return by_store


def aggregate(rows):
    """채널 단위로 집계 (배민배달+배달의민족 → 배달의민족 등 동일 채널 합산)."""
    summary = {}
    total_count = 0
    total_amount = 0
    for r in rows:
        ch = r["channel"]
        if ch not in summary:
            summary[ch] = {"count": 0, "amount": 0}
        summary[ch]["count"] += r["count"]
        summary[ch]["amount"] += r["amount"]
        total_count += r["count"]
        total_amount += r["amount"]
    return summary, total_count, total_amount


def format_won(n):
    return f"{n:,}원"


def diff_html(curr, prev):
    """증감 % HTML 문자열. prev가 0이면 N/A."""
    if not prev:
        return "<span style='color:#999;'>N/A</span>"
    diff = curr - prev
    pct = diff / prev * 100
    if diff > 0:
        return f"<span style='color:#e03131;'>▲ {abs(pct):.1f}%</span>"
    elif diff < 0:
        return f"<span style='color:#1971c2;'>▼ {abs(pct):.1f}%</span>"
    else:
        return "<span style='color:#666;'>━ 0.0%</span>"


def get_ai_analysis(report_type, brand_blocks, grand_cur, grand_prev, cur_range, prev_range):
    """전체 + 브랜드별 + 비교 데이터를 묶어서 Claude AI 분석."""
    period_label = "어제 (전일)" if report_type == "daily" else "지난 7일 (주간)"
    compare_label = "전일" if report_type == "daily" else "전주"

    def pct_str(cur, prev):
        if not prev:
            return "N/A"
        return f"{(cur-prev)/prev*100:+.1f}%"

    brand_sections = []
    for b in brand_blocks:
        lines = []
        for ch, d in sorted(b["cur_summary"].items(), key=lambda x: -x[1]["amount"]):
            ratio = d["amount"] * 100 / b["cur_total_amount"] if b["cur_total_amount"] else 0
            lines.append(f"  - {ch}: {d['count']:,}건, {d['amount']:,}원 ({ratio:.1f}%)")
        brand_sections.append(
            f"[{b['brand']}]\n"
            f"  현재: {b['cur_total_count']:,}건 / {b['cur_total_amount']:,}원\n"
            f"  비교: {b['prev_total_count']:,}건 / {b['prev_total_amount']:,}원\n"
            f"  매출 증감: {pct_str(b['cur_total_amount'], b['prev_total_amount'])}\n"
            + "\n".join(lines)
        )
    brand_text = "\n\n".join(brand_sections)

    prompt = f"""다음은 외식 프랜차이즈 여러 브랜드의 {period_label} 매출과 {compare_label} 대비 비교 데이터입니다.

[기간]
- 현재: {cur_range[0]} ~ {cur_range[1]}
- 비교({compare_label}): {prev_range[0]} ~ {prev_range[1]}

[전체 합계]
- 현재 주문 건수: {grand_cur['count']:,}건 ({compare_label} 대비 {pct_str(grand_cur['count'], grand_prev['count'])})
- 현재 매출액: {grand_cur['amount']:,}원 ({compare_label} 대비 {pct_str(grand_cur['amount'], grand_prev['amount'])})

[브랜드별]
{brand_text}

다음 형식으로 한국어로 간결하게 분석해주세요:

1. **매출 총평** (2~3문장): {compare_label} 대비 흐름과 특이점
2. **브랜드별 비교**: 각 브랜드의 증감과 원인 추정
3. **채널별 인사이트** (2~3개): 배민/쿠팡이츠 등 주요 채널 비중과 시사점
4. **개선 제안** (1~2개): 매출 증대를 위해 고려해볼 점"""

    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1500,
        messages=[{"role": "user", "content": prompt}],
    )
    text = msg.content[0].text
    text = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)
    text = text.replace("\n", "<br>")
    return text


def render_brand_section(b, compare_label):
    """브랜드 1개의 HTML 섹션 (비교 컬럼 포함)."""
    cur_total_count = b["cur_total_count"]
    cur_total_amount = b["cur_total_amount"]
    prev_total_count = b["prev_total_count"]
    prev_total_amount = b["prev_total_amount"]
    avg = cur_total_amount / cur_total_count if cur_total_count else 0

    cur_summary = b["cur_summary"]
    prev_summary = b["prev_summary"]
    channel_rows = ""
    for ch, d in sorted(cur_summary.items(), key=lambda x: -x[1]["amount"]):
        ratio = d["amount"] * 100 / cur_total_amount if cur_total_amount else 0
        prev_amount = prev_summary.get(ch, {}).get("amount", 0)
        channel_rows += f"""
        <tr>
          <td style="padding:6px 12px;">{ch}</td>
          <td style="padding:6px 12px; text-align:right;">{d['count']:,}</td>
          <td style="padding:6px 12px; text-align:right; font-weight:bold;">{format_won(d['amount'])}</td>
          <td style="padding:6px 12px; text-align:right; color:#666;">{ratio:.1f}%</td>
          <td style="padding:6px 12px; text-align:right;">{diff_html(d['amount'], prev_amount)}</td>
        </tr>"""

    # 매장별 TOP 10
    cur_stores = b.get("cur_stores", {})
    prev_stores = b.get("prev_stores", {})
    top10 = sorted(cur_stores.values(), key=lambda x: -x["amount"])[:10]
    top10_rows = ""
    for s in top10:
        prev_amount = 0
        for code, ps in prev_stores.items():
            if ps["store_name"] == s["store_name"]:
                prev_amount = ps["amount"]
                break
        top10_rows += f"""
        <tr>
          <td style="padding:6px 12px;">{s['store_name']}</td>
          <td style="padding:6px 12px; text-align:right;">{s['count']:,}</td>
          <td style="padding:6px 12px; text-align:right; font-weight:bold;">{format_won(s['amount'])}</td>
          <td style="padding:6px 12px; text-align:right;">{diff_html(s['amount'], prev_amount)}</td>
        </tr>"""
    top10_section = ""
    if top10_rows:
        top10_section = f"""
      <h4 style="margin-top:20px; color:#1971c2;">매장 매출 TOP 10</h4>
      <table style="border-collapse:collapse; width:100%; margin-top:8px;">
        <thead>
          <tr style="background:#f1f3f5;">
            <th style="padding:6px 12px; text-align:left;">매장명</th>
            <th style="padding:6px 12px; text-align:right;">건수</th>
            <th style="padding:6px 12px; text-align:right;">매출액</th>
            <th style="padding:6px 12px; text-align:right;">{compare_label} 대비</th>
          </tr>
        </thead>
        <tbody>{top10_rows}</tbody>
      </table>
        """

    # 미운영(매출 0) 매장: 현재 기간 데이터에서 매출 0인 매장 또는, 비교 기간에는 있었는데 현재 사라진 매장
    inactive = []
    cur_names = {s["store_name"] for s in cur_stores.values() if s["amount"] > 0}
    for code, s in cur_stores.items():
        if s["amount"] == 0:
            inactive.append(s["store_name"])
    # 비교 기간에는 매출 있었는데 현재 0/없음
    for code, ps in prev_stores.items():
        if ps["amount"] > 0 and ps["store_name"] not in cur_names:
            if ps["store_name"] not in inactive:
                inactive.append(ps["store_name"])

    inactive_section = ""
    if inactive:
        chips = "".join(
            f'<span style="display:inline-block; padding:4px 10px; margin:3px; background:#fff5f5; border:1px solid #ffc9c9; border-radius:12px; font-size:13px; color:#c92a2a;">{n}</span>'
            for n in sorted(inactive)
        )
        inactive_section = f"""
      <h4 style="margin-top:20px; color:#c92a2a;">미운영 매장 ({len(inactive)}개)</h4>
      <div style="margin-top:8px;">{chips}</div>
        """

    return f"""
      <h3 style="margin-top:32px; padding:8px 12px; background:#e7f5ff; border-left:4px solid #1971c2;">
        [{b['brand']}]
      </h3>
      <table style="border-collapse:collapse; width:100%; margin-top:8px;">
        <tbody>
          <tr style="background:#f8f9fa;">
            <td style="padding:8px 12px; width:40%;">총 주문 건수</td>
            <td style="padding:8px 12px; text-align:right; font-weight:bold;">{cur_total_count:,}건</td>
            <td style="padding:8px 12px; text-align:right;">{diff_html(cur_total_count, prev_total_count)}</td>
          </tr>
          <tr>
            <td style="padding:8px 12px;">총 매출액</td>
            <td style="padding:8px 12px; text-align:right; font-weight:bold; color:#e03131;">{format_won(cur_total_amount)}</td>
            <td style="padding:8px 12px; text-align:right;">{diff_html(cur_total_amount, prev_total_amount)}</td>
          </tr>
          <tr style="background:#f8f9fa;">
            <td style="padding:8px 12px;">평균 객단가</td>
            <td style="padding:8px 12px; text-align:right;">{avg:,.0f}원</td>
            <td></td>
          </tr>
        </tbody>
      </table>

      <h4 style="margin-top:20px; color:#1971c2;">채널별 매출</h4>
      <table style="border-collapse:collapse; width:100%; margin-top:8px;">
        <thead>
          <tr style="background:#f1f3f5;">
            <th style="padding:6px 12px; text-align:left;">채널</th>
            <th style="padding:6px 12px; text-align:right;">건수</th>
            <th style="padding:6px 12px; text-align:right;">매출액</th>
            <th style="padding:6px 12px; text-align:right;">비중</th>
            <th style="padding:6px 12px; text-align:right;">{compare_label} 대비</th>
          </tr>
        </thead>
        <tbody>{channel_rows}</tbody>
      </table>

      {top10_section}

      {inactive_section}
    """


def render_html(report_type, brand_blocks, grand_cur, grand_prev, cur_range, prev_range, ai_text, now_str):
    title_type = "일간" if report_type == "daily" else "주간"
    compare_label = "전일" if report_type == "daily" else "전주"
    grand_avg = grand_cur["amount"] / grand_cur["count"] if grand_cur["count"] else 0

    brand_sections = "\n".join(render_brand_section(b, compare_label) for b in brand_blocks)

    # AI 분석은 ai_text가 있을 때만 섹션 표시 (꺼져 있으면 크레딧 미사용 → 섹션 생략)
    ai_section = ""
    if ai_text:
        ai_section = f"""
      <h3 style="margin-top:32px;">AI 분석</h3>
      <div style="background:#f8f9fa; padding:16px; border-radius:8px; line-height:1.6; font-size:14px;">
        {ai_text}
      </div>
"""

    brand_count = len(brand_blocks)
    cur_range_str = f"{cur_range[0]} ~ {cur_range[1]}" if cur_range[0] != cur_range[1] else f"{cur_range[0]}"
    prev_range_str = f"{prev_range[0]} ~ {prev_range[1]}" if prev_range[0] != prev_range[1] else f"{prev_range[0]}"

    summary_table = f"""
      <h3>전체 합계 ({brand_count}개 브랜드)</h3>
      <p style="margin:4px 0; font-size:13px; color:#666;">
        현재 기간: <strong>{cur_range_str}</strong> /
        비교({compare_label}): <strong>{prev_range_str}</strong>
      </p>
      <table style="border-collapse:collapse; width:100%; margin-top:8px;">
        <tbody>
          <tr style="background:#f8f9fa;">
            <td style="padding:8px 12px; width:40%;">총 주문 건수</td>
            <td style="padding:8px 12px; text-align:right; font-weight:bold; font-size:16px;">{grand_cur['count']:,}건</td>
            <td style="padding:8px 12px; text-align:right;">{diff_html(grand_cur['count'], grand_prev['count'])}</td>
          </tr>
          <tr>
            <td style="padding:8px 12px;">총 매출액</td>
            <td style="padding:8px 12px; text-align:right; font-weight:bold; font-size:16px; color:#e03131;">{format_won(grand_cur['amount'])}</td>
            <td style="padding:8px 12px; text-align:right;">{diff_html(grand_cur['amount'], grand_prev['amount'])}</td>
          </tr>
          <tr style="background:#f8f9fa;">
            <td style="padding:8px 12px;">평균 객단가</td>
            <td style="padding:8px 12px; text-align:right;">{grand_avg:,.0f}원</td>
            <td></td>
          </tr>
        </tbody>
      </table>
    """

    html = f"""
    <html><body style="font-family:Arial,sans-serif; max-width:760px; margin:auto; color:#222;">
      <h2 style="border-bottom:2px solid #333; padding-bottom:8px;">
        매출 리포트 ({title_type})
        <span style="font-size:14px; color:#666;">({now_str} KST)</span>
      </h2>

      {summary_table}

      {brand_sections}

      {ai_section}

      <p style="margin-top:24px; font-size:12px; color:#aaa;">
        본 메일은 자동 발송됩니다. matetech.co.kr 데이터 기반.
      </p>
    </body></html>
    """
    return html


def send_email(subject, html_body):
    sender = os.environ["GMAIL_ADDRESS"]
    password = os.environ["GMAIL_APP_PASSWORD"]
    receiver = os.environ["GMAIL_ADDRESS"]

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = receiver
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender, password)
        server.sendmail(sender, receiver, msg.as_string())


if __name__ == "__main__":
    # 일시 정지 가드: matetech API 전환 작업 진행 중이라 자동/수동 실행 모두 스킵.
    # 다시 켜려면 아래 3줄을 제거하세요.
    if os.environ.get("FORCE_RUN_SALES_REPORT", "").strip().lower() not in ("1", "true", "yes"):
        print("[sales_report] 일시 정지 상태입니다. FORCE_RUN_SALES_REPORT=true 로 강제 실행 가능.")
        raise SystemExit(0)

    report_type = os.environ.get("REPORT_TYPE", "daily").strip().lower()
    if report_type not in ("daily", "weekly"):
        print(f"Unknown REPORT_TYPE: {report_type}, fallback to 'daily'")
        report_type = "daily"

    now_str = datetime.now(KST).strftime("%Y-%m-%d %H:%M")
    title_type = "일간" if report_type == "daily" else "주간"

    accounts = load_accounts()
    if not accounts:
        raise RuntimeError("MATETECH_USERNAME/MATETECH_PASSWORD 환경변수가 없습니다.")

    print(f"[{now_str}] {title_type} 매출 리포트 생성 시작 ({len(accounts)}개 계정)")

    brand_blocks = []
    grand_cur = {"count": 0, "amount": 0}
    grand_prev = {"count": 0, "amount": 0}
    cur_range = prev_range = None

    for acc in accounts:
        print(f"\n--- {acc['brand']} ({acc['username']}) 처리 시작 ---")
        try:
            cur_ch, prev_ch, cur_st, prev_st, cur_range, prev_range = fetch_sales_data(
                acc["username"], acc["password"], report_type
            )
            print(f"엑셀 4개 다운로드 완료 (채널 2개 + 매장 2개)")
        except Exception as e:
            print(f"  데이터 수집 실패: {e}")
            continue

        # 채널별
        cur_rows = parse_excel(cur_ch)
        prev_rows = parse_excel(prev_ch)
        cur_summary, cur_total_count, cur_total_amount = aggregate(cur_rows)
        prev_summary, prev_total_count, prev_total_amount = aggregate(prev_rows)
        print(f"  [채널] 현재 {cur_total_count:,}건/{cur_total_amount:,}원, "
              f"비교 {prev_total_count:,}건/{prev_total_amount:,}원")

        # 매장별 (실패해도 빈 dict로 진행)
        try:
            cur_stores = aggregate_stores(parse_store_excel(cur_st))
            prev_stores = aggregate_stores(parse_store_excel(prev_st))
            print(f"  [매장] 현재 {len(cur_stores)}개 매장, 비교 {len(prev_stores)}개 매장")
        except Exception as e:
            print(f"  매장별 파싱 실패: {e}")
            cur_stores = {}
            prev_stores = {}

        brand_blocks.append({
            "brand": acc["brand"],
            "cur_summary": cur_summary,
            "cur_total_count": cur_total_count,
            "cur_total_amount": cur_total_amount,
            "prev_summary": prev_summary,
            "prev_total_count": prev_total_count,
            "prev_total_amount": prev_total_amount,
            "cur_stores": cur_stores,
            "prev_stores": prev_stores,
        })
        grand_cur["count"] += cur_total_count
        grand_cur["amount"] += cur_total_amount
        grand_prev["count"] += prev_total_count
        grand_prev["amount"] += prev_total_amount

    print(f"\n전체 합계 현재: {grand_cur['count']:,}건 / {grand_cur['amount']:,}원")
    print(f"전체 합계 비교: {grand_prev['count']:,}건 / {grand_prev['amount']:,}원")

    # AI 분석은 기본적으로 꺼져 있어 Claude 크레딧을 쓰지 않는다.
    # 다시 켜려면 환경변수 ENABLE_AI_ANALYSIS=true 로 설정.
    ai_text = None
    if os.environ.get("ENABLE_AI_ANALYSIS", "").strip().lower() in ("1", "true", "on", "yes"):
        try:
            ai_text = get_ai_analysis(
                report_type, brand_blocks, grand_cur, grand_prev, cur_range, prev_range,
            )
            print("AI 분석 완료")
        except Exception as e:
            print(f"AI 분석 실패 (메일은 계속 발송): {e}")
            ai_text = (
                f"<span style='color:#aaa;'>AI 분석을 가져오지 못했습니다.<br>"
                f"({type(e).__name__}: {str(e)[:200]})</span>"
            )
    else:
        print("AI 분석 비활성화됨 (크레딧 미사용). 켜려면 ENABLE_AI_ANALYSIS=true")

    html = render_html(
        report_type, brand_blocks, grand_cur, grand_prev,
        cur_range, prev_range, ai_text, now_str,
    )
    brand_names = " / ".join(b["brand"] for b in brand_blocks)
    subject = f"[매출 리포트 - {title_type}] {brand_names} ({now_str})"
    send_email(subject, html)
    print(f"\n{title_type} 매출 리포트 메일 전송 완료")
