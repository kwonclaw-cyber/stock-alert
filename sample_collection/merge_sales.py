"""matetech 매출(주문경로별/시간대별) + 배민 변경이력 → 통합 타임라인 워크북.

입력:
  - sales_dir: matetech 엑셀들이 있는 폴더
      · 주문경로별_*.xlsx : 일자별 채널 매출 (배민/쿠팡 분리)
      · 시간대별_*.xlsx   : 일자×시간대 매출 (내점/배달/포장)
  - bundle: baemin_console.js 가 만든 변경이력 JSON
출력: 시트 3개(일자별 타임라인 / 시간대별 / 배민 변경이력)짜리 xlsx
사용: python merge_sales.py <sales_dir> <bundle.json> <out.xlsx>
"""

import glob
import json
import os
import re
import sys
import warnings
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

import parse_baemin

warnings.filterwarnings("ignore")
WEEK = ["월", "화", "수", "목", "금", "토", "일"]


def dow(dstr):
    try:
        y, m, d = map(int, dstr.split("-"))
        return WEEK[date(y, m, d).weekday()]
    except Exception:
        return ""


def is_ju(f):  # 주문경로별
    return os.path.basename(f).startswith("#Uc8fc") or "주문경로별" in os.path.basename(f)


def is_si(f):  # 시간대별
    return os.path.basename(f).startswith("#Uc2dc") or "시간대별" in os.path.basename(f)


def parse_channel_daily(files):
    """주문경로별 → {date: {배민:[cnt,amt], 쿠팡:[cnt,amt], 총:[cnt,amt]}}"""
    daily = defaultdict(lambda: {"배민": [0, 0], "쿠팡": [0, 0], "총": [0, 0]})
    for f in files:
        ws = openpyxl.load_workbook(f, data_only=True).active
        for row in ws.iter_rows(min_row=4, values_only=True):
            d, ch = row[0], row[2]
            if not d or d == "합계" or not ch:
                continue
            d = str(d)[:10]
            cnt, amt = int(row[4] or 0), int(row[6] or 0)
            daily[d]["총"][0] += cnt; daily[d]["총"][1] += amt
            if ch == "배달의민족":
                daily[d]["배민"][0] += cnt; daily[d]["배민"][1] += amt
            elif ch == "쿠팡이츠":
                daily[d]["쿠팡"][0] += cnt; daily[d]["쿠팡"][1] += amt
    return daily


def parse_hourly(files):
    """시간대별 → [(date, 시간대, 배달건수, 배달매출, 총건수, 총매출)] 정렬 리스트."""
    rows = []
    for f in files:
        ws = openpyxl.load_workbook(f, data_only=True).active
        cur_date = None
        for row in ws.iter_rows(min_row=4, values_only=True):
            raw_d, slot = row[0], row[1]
            if raw_d and str(raw_d) != "합계":
                m = re.match(r"\s*(\d{1,2})월\s*(\d{1,2})일", str(raw_d))
                if m:
                    mo, dy = int(m.group(1)), int(m.group(2))
                    yr = 2025 if mo == 12 else 2026
                    cur_date = f"{yr}-{mo:02d}-{dy:02d}"
            if not slot or str(slot) == "합계" or not re.match(r"\d", str(slot)):
                continue
            if cur_date is None:
                continue
            tot_cnt = int(row[2] or 0); tot_amt = int(row[4] or 0)
            dlv_cnt = int(row[8] or 0); dlv_amt = int(row[9] or 0)
            rows.append((cur_date, str(slot), dlv_cnt, dlv_amt, tot_cnt, tot_amt))
    rows.sort(key=lambda x: (x[0], x[1]))
    return rows


def change_summary_by_date(chg_rows):
    """변경이력 행 → {date: 요약문자열}, 그리고 운영시간 변경 {date: after}"""
    by_date = defaultdict(list)
    op_hours = {}
    for r in chg_rows:
        d = r["변경일"]
        if not d:
            continue
        gubun, item = r["구분"], r["항목"]
        if gubun == "광고":
            desc = f"광고[{item.split('(')[-1].rstrip(')')}] {r['변경후']}"
        elif gubun == "즉시할인":
            desc = f"즉시할인 {item}"
        elif gubun == "가게":
            desc = item
            if item == "운영시간":
                op_hours[d] = r["변경후"].replace("\n", " / ")
        else:
            desc = f"{gubun} {item}"
        by_date[d].append(desc)
    summary = {d: " | ".join(dict.fromkeys(v)) for d, v in by_date.items()}
    return summary, op_hours


def style_header(ws, ncol):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1971C2")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def build(sales_dir, bundle, out):
    files = glob.glob(os.path.join(sales_dir, "*.xlsx"))
    ju = [f for f in files if is_ju(f)]
    si = [f for f in files if is_si(f)]
    print(f"주문경로별 {len(ju)}개, 시간대별 {len(si)}개")

    daily = parse_channel_daily(ju)
    hourly = parse_hourly(si)
    chg_rows = parse_baemin.parse_bundle(bundle)
    summary, op_hours = change_summary_by_date(chg_rows)

    wb = openpyxl.Workbook()

    # 시트1: 일자별 타임라인
    ws1 = wb.active; ws1.title = "일자별 타임라인"
    cols1 = ["날짜", "요일", "배민매출", "배민건수", "쿠팡매출", "쿠팡건수", "총매출", "총건수", "그날 운영변경(배민)"]
    ws1.append(cols1)
    for d in sorted(daily):
        v = daily[d]
        ws1.append([d, dow(d), v["배민"][1], v["배민"][0], v["쿠팡"][1], v["쿠팡"][0],
                    v["총"][1], v["총"][0], summary.get(d, "")])
    style_header(ws1, len(cols1))
    for col, w in zip("ABCDEFGHI", [12, 6, 13, 9, 13, 9, 13, 9, 70]):
        ws1.column_dimensions[col].width = w
    for r in ws1.iter_rows(min_row=2):
        for c in r[2:8]:
            c.number_format = "#,##0"
        r[8].alignment = Alignment(wrap_text=True, vertical="top")
        if r[8].value:  # 변경 있는 날 강조
            for c in r:
                c.fill = PatternFill("solid", fgColor="FFF3BF")

    # 시트2: 시간대별
    ws2 = wb.create_sheet("시간대별")
    cols2 = ["날짜", "요일", "시간대", "배달건수", "배달매출", "총건수", "총매출", "운영시간변경"]
    ws2.append(cols2)
    for (d, slot, dc, da, tc, ta) in hourly:
        ws2.append([d, dow(d), slot, dc, da, tc, ta, op_hours.get(d, "")])
    style_header(ws2, len(cols2))
    for col, w in zip("ABCDEFGH", [12, 6, 10, 10, 13, 10, 13, 40]):
        ws2.column_dimensions[col].width = w
    for r in ws2.iter_rows(min_row=2):
        for c in (r[4], r[6]):
            c.number_format = "#,##0"
        if r[7].value:
            r[7].fill = PatternFill("solid", fgColor="FFF3BF")

    # 시트3: 배민 변경이력 원본
    ws3 = wb.create_sheet("배민 변경이력")
    cols3 = ["변경일", "변경일시", "구분", "항목", "변경전", "변경후", "작업자", "경로"]
    ws3.append(cols3)
    for r in chg_rows:
        ws3.append([r.get(c, "") for c in cols3])
    style_header(ws3, len(cols3))
    for col, w in zip("ABCDEFGH", [12, 20, 8, 24, 38, 38, 16, 14]):
        ws3.column_dimensions[col].width = w
    for r in ws3.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(out)
    print(f"저장: {out}")
    print(f"  일자별 {ws1.max_row-1}행 / 시간대별 {ws2.max_row-1}행 / 변경이력 {ws3.max_row-1}행")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2], sys.argv[3])
