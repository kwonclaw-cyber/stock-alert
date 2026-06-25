"""매출×변경 통합본 → 변경요인별 매출영향 분석 워크북.

입력: merge_sales.py 가 만든 통합 워크북(.xlsx) — '일자별 타임라인' + '배민 변경이력' 시트 사용.
출력 시트:
  1) 일자별 분석   : 운영시간/할인/우리가게클릭 3개 컬럼 분리 + 전주 동일요일 매출 비교
  2) 주차별 비교   : 주간 매출 vs 전주차 + 그 주 변경요약
  3) 월별 비교     : 월 매출 vs 전월 + 그 달 변경요약
  4) 분석요약     : 변경요인별 매출영향 지표
사용: python analyze.py <통합본.xlsx> [출력.xlsx]
"""

import os
import re
import sys
import warnings
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
WEEK = ["월", "화", "수", "목", "금", "토", "일"]
BLUE = "1971C2"; YEL = "FFF3BF"; RED = "E03131"; GRN = "2F9E44"


def d(s):
    return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()


def won(n):
    return f"{int(round(n)):,}" if n else "0"


def pct(cur, prev):
    if not prev:
        return None
    return (cur - prev) / prev * 100


# ---------- 입력 읽기 ----------
def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["일자별 타임라인"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    sales_name = "배민매출" if "배민매출" in idx else ("배달매출" if "배달매출" in idx else "총매출")
    cnt_name = sales_name.replace("매출", "건수")
    daily = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[idx["날짜"]]:
            continue
        dt = str(r[idx["날짜"]])[:10]
        daily[dt] = {"sales": r[idx.get(sales_name, 0)] or 0, "cnt": r[idx.get(cnt_name, 0)] or 0}

    wc = wb["배민 변경이력"]
    h2 = [c.value for c in wc[1]]; i2 = {h: i for i, h in enumerate(h2)}
    changes = []
    for r in wc.iter_rows(min_row=2, values_only=True):
        if not r[i2["변경일"]]:
            continue
        changes.append({
            "date": str(r[i2["변경일"]])[:10],
            "gubun": r[i2["구분"]], "item": r[i2["항목"]] or "",
            "before": r[i2["변경전"]] or "", "after": r[i2["변경후"]] or "",
        })
    return daily, changes, sales_name


# ---------- 변경 분류 ----------
def bid_of(s):
    m = re.search(r"입찰\s*([\d,]+)\s*원", str(s))
    return int(m.group(1).replace(",", "")) if m else None


def categorize(changes):
    """date -> {운영시간:[..], 할인:[..], 우리가게클릭:[..]}"""
    by = defaultdict(lambda: {"운영시간": [], "할인": [], "우리가게클릭": []})
    for c in changes:
        if c["gubun"] == "가게" and c["item"] == "운영시간":
            by[c["date"]]["운영시간"].append(c["after"].replace("\n", " / "))
        elif c["gubun"] == "즉시할인":
            by[c["date"]]["할인"].append(c["item"])
        elif c["gubun"] == "광고" and str(c["item"]).startswith("우리가게클릭"):
            ba, aa = bid_of(c["before"]), bid_of(c["after"])
            if ba is not None and aa is not None:
                arrow = "▲인상" if aa > ba else ("▼인하" if aa < ba else "유지")
                by[c["date"]]["우리가게클릭"].append(f"입찰 {ba}→{aa}원 ({arrow})")
            else:
                by[c["date"]]["우리가게클릭"].append(c["after"] or c["item"])
    return by


# ---------- 집계 ----------
def weekly(daily):
    agg = defaultdict(lambda: [0, 0])  # monday -> [sales,cnt]
    for dt, v in daily.items():
        mon = d(dt) - timedelta(days=d(dt).weekday())
        agg[mon][0] += v["sales"]; agg[mon][1] += v["cnt"]
    return agg


def monthly(daily):
    agg = defaultdict(lambda: [0, 0])
    for dt, v in daily.items():
        ym = dt[:7]
        agg[ym][0] += v["sales"]; agg[ym][1] += v["cnt"]
    return agg


def cat_summary(cats_in_range):
    """여러 날의 분류 dict 묶음 → '운영시간/할인/광고' 요약 문자열 3개."""
    out = {"운영시간": [], "할인": [], "우리가게클릭": []}
    for c in cats_in_range:
        for k in out:
            out[k] += c[k]
    return {k: " ; ".join(dict.fromkeys(v)) for k, v in out.items()}


# ---------- 워크북 작성 ----------
def hdr_style(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def diff_cell(ws, row, col, p):
    c = ws.cell(row=row, column=col)
    if p is None:
        c.value = "N/A"; c.font = Font(color="999999")
    else:
        c.value = p / 100; c.number_format = "+0.0%;-0.0%"
        c.font = Font(color=RED if p > 0 else (BLUE if p < 0 else "666666"), bold=abs(p) >= 10)


def build(path, out):
    daily, changes, sales_name = load(path)
    cats = categorize(changes)
    days = sorted(daily)
    wb = openpyxl.Workbook()

    # 1) 일자별 분석
    ws = wb.active; ws.title = "일자별 분석"
    cols = ["날짜", "요일", f"{sales_name}", "건수", "운영시간 변경", "할인 변경",
            "우리가게클릭 변경", "전주동일요일 매출", "차이(원)", "차이(%)"]
    ws.append(cols)
    for dt in days:
        v = daily[dt]; ca = cats.get(dt, {"운영시간": [], "할인": [], "우리가게클릭": []})
        prev = daily.get((d(dt) - timedelta(days=7)).isoformat())
        pv = prev["sales"] if prev else None
        diff = (v["sales"] - pv) if pv is not None else None
        ws.append([dt, WEEK[d(dt).weekday()], v["sales"], v["cnt"],
                   " ; ".join(ca["운영시간"]), " ; ".join(ca["할인"]),
                   " ; ".join(ca["우리가게클릭"]),
                   pv if pv is not None else "", diff if diff is not None else "", ""])
        diff_cell(ws, ws.max_row, 10, pct(v["sales"], pv) if pv else None)
    hdr_style(ws)
    for i, w in enumerate([11, 6, 12, 7, 26, 26, 24, 14, 12, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in ws.iter_rows(min_row=2):
        for ci in (3, 8, 9):
            r[ci].number_format = "#,##0"
        for ci in (4, 5, 6):
            r[ci].alignment = Alignment(wrap_text=True, vertical="top")
        if any(r[ci].value for ci in (4, 5, 6)):  # 변경 있는 날 강조
            for c in r:
                c.fill = PatternFill("solid", fgColor=YEL)

    # 2) 주차별 비교
    wk = weekly(daily); wkeys = sorted(wk)
    ws2 = wb.create_sheet("주차별 비교")
    ws2.append(["주차(월~일)", f"{sales_name}", "건수", "전주차 매출", "차이(원)", "차이(%)",
                "운영시간 변경", "할인 변경", "우리가게클릭 변경"])
    for mon in wkeys:
        s, c = wk[mon]
        prev = wk.get(mon - timedelta(days=7))
        pv = prev[0] if prev else None
        # 그 주의 변경 분류
        wk_cats = [cats[dt] for dt in days if mon <= d(dt) <= mon + timedelta(days=6) and dt in cats]
        cs = cat_summary(wk_cats)
        label = f"{mon.isoformat()} ~ {(mon+timedelta(days=6)).isoformat()[5:]}"
        ws2.append([label, s, c, pv if pv else "", (s - pv) if pv else "",
                    "", cs["운영시간"], cs["할인"], cs["우리가게클릭"]])
        diff_cell(ws2, ws2.max_row, 6, pct(s, pv) if pv else None)
    hdr_style(ws2)
    for i, w in enumerate([22, 12, 7, 13, 12, 9, 24, 24, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in ws2.iter_rows(min_row=2):
        for ci in (1, 3, 4):
            r[ci].number_format = "#,##0"
        for ci in (6, 7, 8):
            r[ci].alignment = Alignment(wrap_text=True, vertical="top")
        if any(r[ci].value for ci in (6, 7, 8)):
            for c in r:
                c.fill = PatternFill("solid", fgColor=YEL)

    # 3) 월별 비교
    mo = monthly(daily); mkeys = sorted(mo)
    ws3 = wb.create_sheet("월별 비교")
    ws3.append(["월", f"{sales_name}", "건수", "전월 매출", "차이(원)", "차이(%)",
                "그 달 운영시간", "그 달 할인", "그 달 우리가게클릭"])
    for i, ym in enumerate(mkeys):
        s, c = mo[ym]
        pv = mo[mkeys[i-1]][0] if i > 0 else None
        m_cats = [cats[dt] for dt in days if dt[:7] == ym and dt in cats]
        cs = cat_summary(m_cats)
        ws3.append([ym, s, c, pv if pv else "", (s - pv) if pv else "",
                    "", cs["운영시간"], cs["할인"], cs["우리가게클릭"]])
        diff_cell(ws3, ws3.max_row, 6, pct(s, pv) if pv else None)
    hdr_style(ws3)
    for i, w in enumerate([10, 13, 8, 13, 12, 9, 22, 22, 22], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for r in ws3.iter_rows(min_row=2):
        for ci in (1, 3, 4):
            r[ci].number_format = "#,##0"
        for ci in (6, 7, 8):
            r[ci].alignment = Alignment(wrap_text=True, vertical="top")

    # 4) 분석요약
    ws4 = wb.create_sheet("분석요약")
    insights = analyze_text(daily, cats, days)
    ws4.column_dimensions["A"].width = 100
    ws4["A1"] = "변경요인별 매출영향 분석 (자동 산출)"
    ws4["A1"].font = Font(bold=True, size=13, color=BLUE)
    for i, line in enumerate(insights, 3):
        cell = ws4.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("■"):
            cell.font = Font(bold=True, size=11)

    wb.save(out)
    print(f"저장: {out}")
    for line in insights:
        print(line)


# ---------- 분석 텍스트(지표) ----------
def avg(xs):
    xs = [x for x in xs if x is not None]
    return sum(xs) / len(xs) if xs else None


def analyze_text(daily, cats, days):
    def wow(dt):
        prev = daily.get((d(dt) - timedelta(days=7)).isoformat())
        if not prev or not prev["sales"]:
            return None
        return (daily[dt]["sales"] - prev["sales"]) / prev["sales"] * 100

    op_days = [dt for dt in days if cats.get(dt, {}).get("운영시간")]
    disc_start = [dt for dt in days if any("[시작]" in x for x in cats.get(dt, {}).get("할인", []))]
    disc_end = [dt for dt in days if any("[종료]" in x for x in cats.get(dt, {}).get("할인", []))]
    bid_up = [dt for dt in days if any("▲인상" in x for x in cats.get(dt, {}).get("우리가게클릭", []))]
    bid_dn = [dt for dt in days if any("▼인하" in x for x in cats.get(dt, {}).get("우리가게클릭", []))]
    all_wow = avg([wow(dt) for dt in days])

    def fmt(a):
        return f"{a:+.1f}%" if a is not None else "N/A"

    L = []
    L.append("■ 전주 동일요일 대비 매출 변화(평균) — 변경 유형별")
    L.append(f"  · 전체 일평균(WoW): {fmt(all_wow)}  ← 비교 기준선")
    L.append(f"  · 운영시간 변경일({len(op_days)}일): {fmt(avg([wow(x) for x in op_days]))}")
    L.append(f"  · 즉시할인 '시작'일({len(disc_start)}일): {fmt(avg([wow(x) for x in disc_start]))}")
    L.append(f"  · 즉시할인 '종료'일({len(disc_end)}일): {fmt(avg([wow(x) for x in disc_end]))}")
    L.append(f"  · 우리가게클릭 입찰▲인상일({len(bid_up)}일): {fmt(avg([wow(x) for x in bid_up]))}")
    L.append(f"  · 우리가게클릭 입찰▼인하일({len(bid_dn)}일): {fmt(avg([wow(x) for x in bid_dn]))}")
    L.append("")
    L.append("■ 해석 가이드")
    L.append("  · 기준선(전체 일평균)보다 높으면, 그 변경이 있던 날의 매출이 평소 추세보다 좋았다는 뜻.")
    L.append("  · '시작'일이 +이고 '종료'일이 −면 할인이 매출을 끌어올렸을 가능성.")
    L.append("  · 단, 같은 날 여러 변경이 겹치거나 표본이 적으면 인과가 아닌 상관일 수 있음(주차/월 비교와 함께 보세요).")
    return L


if __name__ == "__main__":
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(src)[0] + "_분석.xlsx"
    build(src, out)
