"""배민 캡처(network/*.json) → 변경이력 표로 정리.

capture.py 로 저장한 network 폴더의 JSON 중에서
- /v1/modify-history/shop          (가게: 운영시간/주문안내/주문금액·배달팁 등)
- /v1/modify-history/instant-discount (즉시할인 시작/종료)
- /v1/modify-history/ad-campaign   (광고: ON/OFF, 입찰·예산)
- .../promotions/history           (메뉴할인)
을 찾아 한 장의 표(CSV / XLSX)로 합친다.

사용:
    python parse_baemin.py <captures 폴더 경로> [출력경로.xlsx]
폴더 경로를 안 주면, 같은 폴더의 captures/ 아래 가장 최근 실행 폴더를 자동 사용.
"""

import csv
import glob
import json
import os
import sys
from datetime import datetime


def _load_json_with_meta(path):
    """첫 줄 '// {meta}' + 본문 JSON 형태를 읽는다."""
    raw = open(path, encoding="utf-8").read()
    first, _, body = raw.partition("\n")
    url = ""
    if first.startswith("//"):
        try:
            url = json.loads(first[2:].strip()).get("url", "")
        except Exception:
            url = first[2:].strip()
    else:
        body = raw
    try:
        return url, json.loads(body)
    except Exception:
        return url, None


def _norm_dt(s):
    """다양한 시간 문자열을 'YYYY-MM-DD HH:MM:SS'로."""
    if not s:
        return ""
    s = str(s).replace("T", " ")
    return s[:19]


def _date_only(s):
    return _norm_dt(s)[:10]


def parse_records(net_dir):
    """network 폴더 내 modify-history / promotion 응답을 표 행 리스트로."""
    rows = []
    seen = set()  # 중복 제거 (id 기준)

    for path in sorted(glob.glob(os.path.join(net_dir, "*.json"))):
        url, data = _load_json_with_meta(path)
        if not isinstance(data, (dict, list)):
            continue

        # 가게 / 즉시할인 / 광고: {content:[...]} 형태
        content = data.get("content") if isinstance(data, dict) else None

        if "modify-history/shop" in url and content:
            for r in content:
                rid = ("shop", r.get("id"))
                if rid in seen:
                    continue
                seen.add(rid)
                rows.append({
                    "구분": "가게",
                    "항목": _shop_item_label(r.get("changedItem"), r.get("subDivision")),
                    "변경전": (r.get("beforeState") or "").strip(),
                    "변경후": (r.get("afterState") or "").strip(),
                    "변경일시": _norm_dt(r.get("modified") or r.get("createdAt")),
                    "변경일": _date_only(r.get("modified") or r.get("createdAt")),
                    "작업자": r.get("modifyId") or "",
                    "경로": r.get("modifyChannel") or "",
                    "_raw_item": r.get("changedItem") or "",
                })

        elif "modify-history/instant-discount" in url and content:
            for r in content:
                rid = ("disc", r.get("id"))
                if rid in seen:
                    continue
                seen.add(rid)
                rows.append({
                    "구분": "즉시할인",
                    "항목": r.get("name") or r.get("type") or "",
                    "변경전": "",
                    "변경후": (r.get("description") or "").strip(),
                    "변경일시": _norm_dt(r.get("modifiedAt")),
                    "변경일": _date_only(r.get("modifiedAt")),
                    "작업자": r.get("modifiedBy") or "",
                    "경로": "",
                    "_raw_item": r.get("type") or "",
                })

        elif "modify-history/ad-campaign" in url and content:
            for r in content:
                # 광고는 id가 null인 경우가 있어 시간+타입으로 키 생성
                rid = ("ad", r.get("adCampaignId"), r.get("historyType"),
                       r.get("createdAt"), r.get("beforeValue"), r.get("afterValue"))
                if rid in seen:
                    continue
                seen.add(rid)
                rows.append({
                    "구분": "광고",
                    "항목": f"{r.get('adKindTitle','')} ({_ad_type_label(r.get('historyType'))})".strip(),
                    "변경전": _fmt_ad_value(r.get("beforeValue")),
                    "변경후": _fmt_ad_value(r.get("afterValue")),
                    "변경일시": _norm_dt(r.get("createdAt")),
                    "변경일": _date_only(r.get("createdAt")),
                    "작업자": r.get("createdById") or "",
                    "경로": "",
                    "_raw_item": r.get("historyType") or "",
                })

        elif "promotions/history" in url and content:
            for r in content:
                rid = ("promo", r.get("id") or json.dumps(r, ensure_ascii=False)[:50])
                if rid in seen:
                    continue
                seen.add(rid)
                rows.append({
                    "구분": "메뉴할인",
                    "항목": r.get("name") or r.get("promotionName") or r.get("title") or "",
                    "변경전": "",
                    "변경후": json.dumps(r, ensure_ascii=False),
                    "변경일시": _norm_dt(r.get("modifiedAt") or r.get("createdAt") or r.get("updatedAt")),
                    "변경일": _date_only(r.get("modifiedAt") or r.get("createdAt") or r.get("updatedAt")),
                    "작업자": r.get("modifiedBy") or "",
                    "경로": "",
                    "_raw_item": "MENU_PROMOTION",
                })

    rows.sort(key=lambda x: x["변경일시"])
    return rows


_SHOP_LABELS = {
    "SHOP_OPERATION_HOUR_MODIFY": "운영시간",
    "SHOP_MENU_INFO_MENU_INTRO_MODIFY": "주문안내",
    "SHOP_DELIVERY_RULES_MODIFY": "주문금액·배달팁",
    "SHOP_DELIVERY_RULES_RESERVATION_MODIFY": "주문금액·배달팁(예약)",
    "SHOP_REGION_DELIVERY_FEE_MODIFY": "지역 할증 배달팁",
}


def _shop_item_label(changed_item, sub):
    return _SHOP_LABELS.get(changed_item, changed_item or sub or "기타")


_AD_TYPE_LABELS = {
    "LISTING_INVENTORY_DISPLAY_PAUSE": "노출 ON/OFF",
    "CPC_BUDGET": "입찰·예산",
}


def _ad_type_label(t):
    return _AD_TYPE_LABELS.get(t, t or "")


def _fmt_ad_value(v):
    """광고 값은 가끔 JSON 문자열(cpc bid/budget) → 읽기 쉽게."""
    if v is None:
        return ""
    s = str(v)
    if s.startswith("{"):
        try:
            o = json.loads(s)
            cpc = o.get("cpc", o)
            parts = []
            if "bid" in cpc:
                parts.append(f"입찰 {cpc['bid']:,}원")
            if "maxBid" in cpc:
                parts.append(f"최대입찰 {cpc['maxBid']:,}원")
            if "monthlyBudget" in cpc:
                parts.append(f"월예산 {cpc['monthlyBudget']:,}원")
            if "autoBidding" in cpc:
                parts.append("자동입찰" if cpc["autoBidding"] else "수동입찰")
            return ", ".join(parts) if parts else s
        except Exception:
            return s
    return s


def write_outputs(rows, out_base):
    cols = ["변경일", "변경일시", "구분", "항목", "변경전", "변경후", "작업자", "경로"]
    csv_path = out_base + ".csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"CSV 저장: {csv_path}  ({len(rows)}행)")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "변경이력"
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1971C2")
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        widths = [12, 20, 10, 26, 40, 40, 16, 14]
        for i, wdt in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        xlsx_path = out_base + ".xlsx"
        wb.save(xlsx_path)
        print(f"XLSX 저장: {xlsx_path}")
        return xlsx_path
    except Exception as e:
        print(f"XLSX 저장 건너뜀: {e}")
        return csv_path


def find_latest_captures():
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "captures")
    runs = sorted(glob.glob(os.path.join(base, "*")))
    return runs[-1] if runs else None


def main():
    target = sys.argv[1] if len(sys.argv) > 1 else find_latest_captures()
    if not target or not os.path.isdir(target):
        print("captures 폴더 경로를 인자로 주세요. 예: python parse_baemin.py captures/20260609_092839")
        sys.exit(1)
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(target, "변경이력_정리")
    out = out[:-5] if out.endswith(".xlsx") else out

    # 모든 매장 폴더의 network 디렉터리를 훑는다
    net_dirs = glob.glob(os.path.join(target, "**", "network"), recursive=True)
    if not net_dirs:
        print(f"network 폴더를 못 찾음: {target}")
        sys.exit(1)

    all_rows = []
    for nd in net_dirs:
        store = os.path.basename(os.path.dirname(nd))
        rows = parse_records(nd)
        for r in rows:
            r["매장폴더"] = store
        print(f"  {store}: {len(rows)}건")
        all_rows.extend(rows)

    write_outputs(all_rows, out)


if __name__ == "__main__":
    main()
